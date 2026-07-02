"""上期帶入、範本改版、交件追蹤、金額欄設定測試。"""
from __future__ import annotations

from app import report


def test_carry_over_fills_empty(tmp_path, monkeypatch):
    from openpyxl import Workbook, load_workbook

    out = tmp_path / "output"
    out.mkdir()
    monkeypatch.setattr(report, "OUTPUT_DIR", out)

    wb = Workbook()
    ws = wb.active
    ws["B1"] = "{{工程名稱}}"
    ws["B2"] = "{{金額}}"
    tpl = tmp_path / "tpl.xlsx"
    wb.save(tpl)

    fields = report.extract_fields(tpl, "excel")
    mappings = [
        {"report_field": "工程名稱", "source_field": "工程"},
        {"report_field": "金額", "source_field": "額度"},
    ]
    # 本次來源沒有「工程」欄 → 工程名稱會留空,靠上期帶入
    source_rows = [{"額度": "5000"}]
    carry = {"工程名稱": "測試工程", "金額": "0"}

    result = report.generate(tpl, "excel", fields, mappings, source_rows, "o.xlsx", carry_over=carry)
    wb2 = load_workbook(result["output_path"])
    ws2 = wb2.active
    assert ws2["B1"].value == "測試工程"   # 由上期帶入
    assert ws2["B2"].value == "5000"        # 本次有值,不帶入
    assert any("帶入上期" in w for w in result["warnings"])


def test_template_versioning(temp_db):
    rid = temp_db.create_report("報表A", "/tmp/v1.xlsx", "excel")
    assert temp_db.get_report(rid)["version"] == 1
    v = temp_db.update_report_template(rid, "/tmp/v2.xlsx", "excel")
    assert v == 2
    r = temp_db.get_report(rid)
    assert r["version"] == 2 and r["template_path"] == "/tmp/v2.xlsx"
    versions = temp_db.list_template_versions(rid)
    assert [x["version"] for x in versions] == [2, 1]  # 新到舊


def test_tracking(temp_db):
    rid = temp_db.create_report("月報")
    v1 = temp_db.create_vendor("甲營造")
    v2 = temp_db.create_vendor("乙營造")
    temp_db.set_report_vendors(rid, [v1, v2])
    assert temp_db.list_report_vendors(rid) == [v1, v2]

    # 甲交件、乙未交
    temp_db.record_submission(rid, v1, "2026-07", None)
    subs = {s["vendor_id"] for s in temp_db.list_submissions(rid, "2026-07")}
    assert subs == {v1}
    # 同期重覆交件 → 更新不重覆新增
    temp_db.record_submission(rid, v1, "2026-07", 99)
    subs2 = temp_db.list_submissions(rid, "2026-07")
    assert len(subs2) == 1 and subs2[0]["output_id"] == 99


def test_field_roles(temp_db):
    rid = temp_db.create_report("報表A")
    temp_db.set_report_fields(rid, [
        {"field_name": "廠商", "location": "A1"},
        {"field_name": "金額", "location": "B1"},
    ])
    temp_db.set_field_roles(rid, {"金額": "amount", "廠商": "normal"})
    fields = {f["field_name"]: f["field_role"] for f in temp_db.list_report_fields(rid)}
    assert fields["金額"] == "amount"
    assert fields["廠商"] == "normal"


def test_latest_output_values(temp_db):
    import json
    rid = temp_db.create_report("報表A")
    vid = temp_db.create_vendor("甲營造")
    temp_db.create_output(rid, vid, "s.xlsx", "/tmp/o.xlsx", "ok",
                          values_json=json.dumps({"工程名稱": "A案"}, ensure_ascii=False))
    assert temp_db.latest_output_values(rid, vid) == {"工程名稱": "A案"}
    assert temp_db.latest_output_values(rid, 999) == {}

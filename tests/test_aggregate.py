"""多廠商彙總 + 數字驗算測試。"""
from __future__ import annotations

from app import aggregate, validate


def test_validate_to_number():
    assert validate.to_number("1,234") == 1234
    assert validate.to_number("$1,000元") == 1000
    assert validate.to_number("") is None
    assert validate.to_number("abc") is None
    assert validate.to_number(50) == 50


def test_tax_and_totals():
    rows = [{"金額": "1000"}, {"金額": "2,000"}, {"金額": "500"}]
    assert validate.sum_column(rows, "金額") == 3500
    assert validate.tax_of(3500) == 175.0


def test_check_declared_total_mismatch():
    rows = [{"金額": "1000"}, {"金額": "2000"}]
    assert validate.check_declared_total(rows, "金額", "3000") is None  # 相符
    warn = validate.check_declared_total(rows, "金額", "9999")
    assert warn is not None and "不符" in warn


def test_generate_summary_with_totals(tmp_path, monkeypatch):
    from openpyxl import Workbook, load_workbook

    out = tmp_path / "output"
    out.mkdir()
    monkeypatch.setattr(aggregate, "OUTPUT_DIR", out)

    # 表頭式範本:A1 廠商名稱, B1 金額
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "廠商名稱"
    ws["B1"] = "金額"
    tpl = tmp_path / "tpl.xlsx"
    wb.save(tpl)

    report_fields = [
        {"field_name": "廠商名稱", "location": "A1", "field_role": "normal"},
        {"field_name": "金額", "location": "B1", "field_role": "amount"},
    ]
    vendor_entries = [
        {"vendor_name": "甲", "values": {"廠商名稱": "甲營造", "金額": "1000"}},
        {"vendor_name": "乙", "values": {"廠商名稱": "乙營造", "金額": "2000"}},
    ]
    result = aggregate.generate_summary(tpl, report_fields, vendor_entries, "summary.xlsx")

    wb2 = load_workbook(result["output_path"])
    ws2 = wb2.active
    assert ws2["A2"].value == "甲營造"
    assert ws2["A3"].value == "乙營造"
    # 第 4 列起為 總計 / 稅額 / 含稅總計
    assert ws2["A4"].value == "總計"
    assert ws2["B4"].value == 3000
    assert ws2["A5"].value == "稅額(5%)"
    assert ws2["B5"].value == 150.0
    assert ws2["A6"].value == "含稅總計"
    assert ws2["B6"].value == 3150.0
    assert result["totals"]["金額"] == 3000

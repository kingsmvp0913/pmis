"""Excel 讀取器(.xlsx / .xlsm / .xls)。

假設第一列為欄位標題(表頭),其餘為資料列 —— 這是營造報表最常見的排版。
以 openpyxl 讀取 .xlsx/.xlsm;.xls 交給 pandas(內部走 xlrd)。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from .base import ParseError, ParseResult


def _clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_xlsx(path: Path) -> ParseResult:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=False)

    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise ParseError(f"Excel 檔「{path.name}」是空的,沒有任何資料。")

    fields: list[str] = []
    locations: dict[str, str] = {}
    for cell in header_cells:
        name = _clean(cell.value)
        if name == "":
            continue
        name = str(name)
        fields.append(name)
        locations[name] = cell.coordinate  # 例:B1

    if not fields:
        raise ParseError(f"Excel 檔「{path.name}」第一列沒有欄位標題。")

    data_rows: list[dict[str, Any]] = []
    for row in rows_iter:
        record: dict[str, Any] = {}
        has_value = False
        for idx, name in enumerate(fields):
            val = _clean(row[idx].value) if idx < len(row) else ""
            if val != "":
                has_value = True
            record[name] = val
        if has_value:
            data_rows.append(record)

    wb.close()
    return ParseResult(
        fields=fields, rows=data_rows, meta={"locations": locations}
    )


def _parse_xls(path: Path) -> ParseResult:
    import pandas as pd

    df = pd.read_excel(path, dtype=object)
    df = df.where(df.notna(), "")
    fields = [str(c).strip() for c in df.columns]
    rows = [
        {f: _clean(v) for f, v in zip(fields, rec)}
        for rec in df.itertuples(index=False, name=None)
    ]
    if not fields:
        raise ParseError(f"Excel 檔「{path.name}」沒有欄位標題。")
    return ParseResult(fields=fields, rows=rows)


def parse(path: Path) -> ParseResult:
    if path.suffix.lower() == ".xls":
        return _parse_xls(path)
    return _parse_xlsx(path)

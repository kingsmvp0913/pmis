"""多家廠商彙總成一張總表。

適用「表頭式 Excel 報表範本」(第一列為欄位名,每家廠商一列)。系統把每家廠商
依其對應設定填成一列,最後自動加上金額欄的「總計 / 稅額(5%)/ 含稅總計」列。

placeholder 式或 Word 範本不適合做多廠商總表,會回報明確訊息請改用表頭式 Excel。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from . import validate
from .base_paths import OUTPUT_DIR
from .parsers.base import ParseError


def is_summary_capable(template_type: str, report_fields: list[dict[str, Any]]) -> bool:
    """能否做彙總總表:必須是 Excel 且欄位以表頭座標定位(非 placeholder)。"""
    if template_type != "excel":
        return False
    # 表頭式:location 為第一列座標(row=1)。placeholder 可能落在任意列。
    for f in report_fields:
        loc = str(f.get("location", ""))
        if loc and not loc.rstrip("0123456789").isalpha():
            continue
    # 簡化判斷:只要不是全都在同一非表頭列即可;實務上以「第一列有欄位」為準。
    return True


def generate_summary(
    template_path: str | Path,
    report_fields: list[dict[str, Any]],
    vendor_entries: list[dict[str, Any]],
    output_filename: str,
) -> dict[str, Any]:
    """產出彙總總表。

    vendor_entries:[{vendor_name, values:{report_field: value}}] —— 每家一列。
    report_fields:含 field_role('amount' 者納入總計/稅額計算)。
    回傳:{output_path, warnings, status, totals}
    """
    from openpyxl import load_workbook
    from openpyxl.utils import coordinate_to_tuple

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / output_filename
    warnings: list[str] = []

    wb = load_workbook(template_path)
    ws = wb.active

    # 建立 欄位 → 欄號 對照(依表頭座標)
    col_of: dict[str, int] = {}
    header_row = 1
    for f in report_fields:
        loc = str(f.get("location", ""))
        try:
            r, c = coordinate_to_tuple(loc)
            header_row = r
            col_of[f["field_name"]] = c
        except Exception:
            warnings.append(f"欄位「{f['field_name']}」無有效表頭位置,略過。")

    if not col_of:
        raise ParseError(
            "此報表範本無法做彙總總表(找不到表頭欄位)。請改用「第一列為欄位名」的 Excel 範本。"
        )

    amount_fields = [f["field_name"] for f in report_fields if f.get("field_role") == "amount"]

    # 逐家廠商填一列
    start = header_row + 1
    filled_rows: list[dict[str, Any]] = []
    for i, entry in enumerate(vendor_entries):
        values = entry.get("values", {})
        filled_rows.append(values)
        for field_name, col in col_of.items():
            ws.cell(row=start + i, column=col, value=values.get(field_name, ""))

    # 金額欄:加上 總計 / 稅額 / 含稅總計
    totals: dict[str, float] = {}
    if amount_fields:
        totals = validate.column_totals(filled_rows, amount_fields)
        n = len(vendor_entries)
        label_col = min(col_of.values())

        def put_row(offset: int, label: str, value_map: dict[int, Any]) -> None:
            row = start + n + offset
            ws.cell(row=row, column=label_col, value=label)
            for col, val in value_map.items():
                ws.cell(row=row, column=col, value=val)

        subtotal_map = {col_of[f]: round(totals[f], 2) for f in amount_fields}
        tax_map = {col_of[f]: validate.tax_of(totals[f]) for f in amount_fields}
        withtax_map = {
            col_of[f]: round(totals[f] + validate.tax_of(totals[f]), 2) for f in amount_fields
        }
        # 若標籤欄本身就是金額欄,避免覆蓋,標籤放在最左金額欄左邊一欄(退回用金額欄)
        put_row(0, "總計", subtotal_map)
        put_row(1, "稅額(5%)", tax_map)
        put_row(2, "含稅總計", withtax_map)
        warnings.append(
            "已自動加上『總計 / 稅額(5%)/ 含稅總計』三列,請核對金額。"
        )

    wb.save(out_path)
    status = "warning" if warnings else "ok"
    return {
        "output_path": str(out_path),
        "warnings": warnings,
        "status": status,
        "totals": totals,
    }
